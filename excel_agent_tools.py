"""
Excel Agent Tools

This module contains the tool definitions for the Excel Agent.
These tools allow the agent to manipulate Excel files through a chat interface using openpyxl.
"""
import os
import json
import base64
import tempfile
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, LineChart, PieChart
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
import matplotlib.pyplot as plt
from io import BytesIO
from typing import Optional, Any, Dict, List
from langchain.callbacks.manager import CallbackManagerForToolRun
from langchain.tools import BaseTool
import chainlit as cl
import shutil
import copy
from chainlit import run_sync

# Global variables to store Excel workbooks and the current file ID
# These are now just placeholders - actual data is stored in Chainlit user session
EXCEL_FILES = {}
CURRENT_FILE_ID = None

# Function to get Excel files from Chainlit session
def get_excel_files():
    """Get Excel files from Chainlit user session."""
    files = cl.user_session.get("excel_files")
    if files is None:
        files = {}
        cl.user_session.set("excel_files", files)
    return files

# Function to get current file ID from Chainlit session
def get_current_file_id():
    """Get current file ID from Chainlit user session."""
    return cl.user_session.get("current_file_id")

# Function to set current file ID in Chainlit session
def set_current_file_id(file_id):
    """Set current file ID in Chainlit user session."""
    cl.user_session.set("current_file_id", file_id)

# Function to save Excel files metadata to a JSON file
def save_excel_files_metadata(excel_files_dir, metadata_file):
    """Save the Excel files metadata to a JSON file."""
    # Get Excel files from Chainlit session
    excel_files = get_excel_files()
    
    # Convert workbooks to None for JSON serialization
    metadata = {}
    for file_id, file_info in excel_files.items():
        metadata[file_id] = {
            "file_path": file_info["file_path"],
            "workbook": None  # We don't save the workbook itself
        }
    
    with open(metadata_file, "w") as f:
        json.dump(metadata, f)

# Function to load Excel files metadata from the JSON file
def load_excel_files_metadata(excel_files_dir, metadata_file):
    """Load the Excel files metadata from the JSON file and load the Excel files."""
    # Get Excel files dictionary from Chainlit session
    excel_files = get_excel_files()
    
    if os.path.exists(metadata_file):
        with open(metadata_file, "r") as f:
            metadata = json.load(f)
        
        # Load the Excel files
        for file_id, file_info in metadata.items():
            file_path = file_info["file_path"]
            if os.path.exists(file_path):
                try:
                    wb = openpyxl.load_workbook(file_path)
                    excel_files[file_id] = {
                        "file_path": file_path,
                        "workbook": wb
                    }
                except Exception as e:
                    print(f"Error loading Excel file {file_path}: {str(e)}")
        
        # Update the Excel files in Chainlit session
        cl.user_session.set("excel_files", excel_files)

# Function to clear all Excel files and metadata
def clear_excel_files(excel_files_dir, metadata_file):
    """Clear all Excel files and reset the metadata."""
    # Remove all files in the Excel files directory
    for file_name in os.listdir(excel_files_dir):
        file_path = os.path.join(excel_files_dir, file_name)
        if os.path.isfile(file_path) and file_name.endswith(('.xlsx', '.xls')):
            try:
                os.remove(file_path)
            except Exception as e:
                print(f"Error removing file {file_path}: {str(e)}")
    
    # Reset the metadata in Chainlit session
    cl.user_session.set("excel_files", {})
    cl.user_session.set("current_file_id", None)
    
    # Save empty metadata to file
    save_excel_files_metadata(excel_files_dir, metadata_file)
    return "All Excel files have been cleared."

# Helper function to get workbook info
def get_workbook_info(wb):
    """Get information about an openpyxl workbook."""
    info = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        # Find the data range
        min_row, min_col, max_row, max_col = 1, 1, 1, 1
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    max_row = max(max_row, cell.row)
                    max_col = max(max_col, cell.column)
        
        # Get column headers if they exist
        headers = []
        if max_row > 0 and max_col > 0:
            for col in range(1, max_col + 1):
                cell_value = sheet.cell(row=1, column=col).value
                headers.append(str(cell_value) if cell_value is not None else f"Column {get_column_letter(col)}")
        
        info.append({
            "sheet_name": sheet_name,
            "dimensions": f"{max_row} rows × {max_col} columns",
            "headers": headers
        })
    
    return info

def remove_images_from_workbook(wb):
    """Remove all images from an Excel workbook to prevent I/O errors when saving."""
    print("Removing images from workbook...")
    
    # Track if any images were found
    images_found = False
    
    # Process each sheet in the workbook
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        # Check if the sheet has images
        if hasattr(sheet, '_images'):
            if sheet._images:
                print(f"Found {len(sheet._images)} images in sheet '{sheet_name}'")
                images_found = True
                # Clear the images
                sheet._images = []
        
        # Check for drawings (another way images can be stored)
        if hasattr(sheet, '_drawing'):
            if sheet._drawing is not None:
                print(f"Found drawings in sheet '{sheet_name}'")
                images_found = True
                # Remove the drawing
                sheet._drawing = None
    
    if not images_found:
        print("No images found in workbook")
    else:
        print("Successfully removed all images from workbook")
    
    return wb

# Define the Excel Agent tools

class ListExcelFilesTool(BaseTool):
    name: str = "list_excel_files"
    description: str = "List all Excel files that have been loaded."
    
    def _run(self, tool_input: str = None, run_manager: Optional[CallbackManagerForToolRun] = None) -> str:
        """List all Excel files that have been loaded."""
        # Get Excel files from Chainlit session
        excel_files = get_excel_files()
        
        if not excel_files:
            return "No Excel files have been loaded. Please upload or load an Excel file first."
        
        result = "Loaded Excel files:\n\n"
        for file_id, file_info in excel_files.items():
            file_path = file_info["file_path"]
            wb = file_info["workbook"]
            wb_info = get_workbook_info(wb)
            
            result += f"ID: {file_id}\n"
            result += f"Path: {file_path}\n"
            result += f"Sheets: {', '.join(wb.sheetnames)}\n"
            
            for sheet_info in wb_info:
                result += f"  - {sheet_info['sheet_name']}: {sheet_info['dimensions']}\n"
                if sheet_info['headers']:
                    result += f"    Headers: {', '.join(sheet_info['headers'][:10])}"
                    if len(sheet_info['headers']) > 10:
                        result += f" and {len(sheet_info['headers']) - 10} more"
                    result += "\n"
            
            result += "\n"
        
        return result

class LoadExcelFileTool(BaseTool):
    name: str = "load_excel_file"
    description: str = "Load an Excel file into memory. Input should be the file path."
    
    def _run(self, file_path: str, run_manager: Optional[CallbackManagerForToolRun] = None) -> str:
        """Load an Excel file into memory."""
        # Get Excel files from Chainlit session
        excel_files = get_excel_files()
        
        try:
            # Check if the file exists
            if not os.path.exists(file_path):
                return f"Error: File '{file_path}' not found."
            
            # Load the Excel file
            wb = openpyxl.load_workbook(file_path)
            wb = remove_images_from_workbook(wb)

            # Generate a file ID
            file_id = os.path.basename(file_path)
            
            # Store the file information in Chainlit session
            excel_files[file_id] = {
                "file_path": file_path,
                "workbook": wb
            }
            cl.user_session.set("excel_files", excel_files)
            
            # Set the current file ID in Chainlit session
            set_current_file_id(file_id)
            
            # Save the metadata
            save_excel_files_metadata(os.path.dirname(file_path), os.path.join(os.path.dirname(file_path), "excel_files_metadata.json"))
            
            # Get workbook info
            wb_info = get_workbook_info(wb)
            
            result = f"Excel file '{file_path}' loaded successfully.\n"
            result += f"Sheets: {', '.join(wb.sheetnames)}\n\n"
            
            for sheet_info in wb_info:
                result += f"Sheet: {sheet_info['sheet_name']}\n"
                result += f"Dimensions: {sheet_info['dimensions']}\n"
                if sheet_info['headers']:
                    result += f"Headers: {', '.join(sheet_info['headers'][:10])}"
                    if len(sheet_info['headers']) > 10:
                        result += f" and {len(sheet_info['headers']) - 10} more"
                    result += "\n\n"
            
            return result
        
        except Exception as e:
            return f"Error loading Excel file: {str(e)}"

class GetExcelInfoTool(BaseTool):
    name: str = "get_excel_info"
    description: str = "Get information about a loaded Excel file. Input should be the file ID or leave empty for the current file."
    
    def _run(self, file_id: str = None, run_manager: Optional[CallbackManagerForToolRun] = None) -> str:
        """Get information about a loaded Excel file."""
        # Get Excel files and current file ID from Chainlit session
        excel_files = get_excel_files()
        current_file_id = get_current_file_id()
        
        if file_id is None or file_id.strip() == "":
            file_id = current_file_id
        
        if file_id is None:
            return "No Excel file is currently selected. Use 'load_excel_file' to load a file."
        
        if file_id not in excel_files:
            return f"Error: Excel file with ID {file_id} not found."
        
        wb = excel_files[file_id]["workbook"]
        file_path = excel_files[file_id]["file_path"]
        
        # Get basic information
        info = f"File ID: {file_id}\n"
        info += f"File Path: {file_path}\n"
        info += f"Sheets: {', '.join(wb.sheetnames)}\n\n"
        
        # Get detailed sheet information
        wb_info = get_workbook_info(wb)
        
        for sheet_info in wb_info:
            info += f"Sheet: {sheet_info['sheet_name']}\n"
            info += f"Dimensions: {sheet_info['dimensions']}\n"
            
            if sheet_info['headers']:
                info += f"Headers: {', '.join(sheet_info['headers'])}\n"
            
            # Get a sample of the data
            sheet = wb[sheet_info['sheet_name']]
            info += "\nData Sample (first 5 rows):\n"
            
            for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=6, values_only=True)):
                if row_idx == 0:
                    info += "| " + " | ".join([str(cell) if cell is not None else "" for cell in row]) + " |\n"
                    info += "|" + "|".join(["----" for _ in row]) + "|\n"
                else:
                    info += "| " + " | ".join([str(cell) if cell is not None else "" for cell in row]) + " |\n"
            
            info += "\n"
        
        return info

class DownloadExcelFileTool(BaseTool):
    name: str = "download_excel_file"
    description: str = "Download the current Excel file to your computer."
    
    def _run(self, file_name: str = None, run_manager: Optional[CallbackManagerForToolRun] = None) -> str:
        """Download the current Excel file."""
        # Get Excel files and current file ID from the session
        excel_files = get_excel_files()
        current_file_id = get_current_file_id()
        
        if current_file_id is None:
            return "No Excel file is currently selected. Use 'load_excel_file' to load a file."
        
        if current_file_id not in excel_files:
            return f"Error: Excel file with ID {current_file_id} not found."
        
        # Get the workbook and file path
        wb = excel_files[current_file_id]['workbook']
        file_path = excel_files[current_file_id]['file_path']
        
        # If no file name is provided, use the original file name
        if not file_name:
            file_name = os.path.basename(file_path)
            
        if not file_name.endswith('.xlsx'):
            file_name += '.xlsx'
        
        # Base filename without extension for creating variant names
        base_name = os.path.splitext(file_name)[0]
        
        # Try each approach in sequence, stopping after the first success
        file_data = None
        method_name = ""
        download_filename = ""
        
        try:
            # APPROACH 1: Save directly to BytesIO
            try:
                buffer = BytesIO()
                wb.save(buffer)
                
                # Get the value from the buffer
                buffer.seek(0)
                file_data = buffer.getvalue()
                method_name = "Direct BytesIO Save"
                download_filename = f"{base_name}.xlsx"
                
                # Save a debug copy to disk for inspection (optional)
                debug_path = os.path.join(os.path.dirname(file_path), "debug_" + base_name + "_bytesio.xlsx")
                with open(debug_path, 'wb') as f:
                    f.write(file_data)
                
            except Exception as approach1_error:
                # If first approach fails, try the second approach
                print(f"Approach 1 (BytesIO) failed: {str(approach1_error)}")
                
                # APPROACH 2: Save to temporary file first, then read back
                try:
                    import tempfile
                    
                    # Create a temporary file
                    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
                        temp_path = temp_file.name
                    
                    # Save the workbook to the temporary file
                    wb.save(temp_path)
                    
                    # Read the file back
                    with open(temp_path, 'rb') as f:
                        file_data = f.read()
                    
                    method_name = "Temporary File Save"
                    download_filename = f"{base_name}.xlsx"
                    
                    # Save a debug copy (optional)
                    debug_path = os.path.join(os.path.dirname(file_path), "debug_" + base_name + "_tempfile.xlsx")
                    with open(debug_path, 'wb') as f:
                        f.write(file_data)
                    
                    # Clean up the temporary file
                    os.unlink(temp_path)
                    
                except Exception as approach2_error:
                    # If second approach fails, try the third approach
                    print(f"Approach 2 (Temporary File) failed: {str(approach2_error)}")
                    
                    # APPROACH 3: Direct file read
                    try:
                        # Read the file directly from disk
                        with open(file_path, 'rb') as f:
                            file_data = f.read()
                        
                        method_name = "Direct File Read"
                        download_filename = f"{base_name}.xlsx"
                        
                        # Save a debug copy (optional)
                        debug_path = os.path.join(os.path.dirname(file_path), "debug_" + base_name + "_directread.xlsx")
                        with open(debug_path, 'wb') as f:
                            f.write(file_data)
                        
                    except Exception as approach3_error:
                        # All approaches failed
                        raise Exception(f"All approaches failed: BytesIO: {str(approach1_error)}, Temp File: {str(approach2_error)}, Direct Read: {str(approach3_error)}")
            
            finally:
                # Ensure BytesIO is closed if it was created
                if 'buffer' in locals():
                    buffer.close()
            
            # If we got here, one of the approaches succeeded
            if file_data is None:
                raise Exception("Failed to generate Excel file data through any method")
            
            # Create a temporary async function to send the file
            async def send_file_async():
                # Create file element with raw binary data (no base64 encoding)
                file_element = cl.File(
                    name=download_filename, 
                    content=file_data, 
                    display="inline", 
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
                await cl.Message(
                    content=f"Here is your Excel file (using {method_name}): {download_filename}", 
                    elements=[file_element]
                ).send()
            
            # Run the async function in the sync context
            run_sync(send_file_async())
            
            return f"Excel file '{download_filename}' has been sent for download using {method_name}."
            
        except Exception as e:
            # Fallback method - direct file read with minimal processing
            try:
                # Check if the file exists
                if not os.path.exists(file_path):
                    return f"Error: File {file_path} does not exist on disk."
                
                # Read the file directly
                with open(file_path, 'rb') as f:
                    file_data = f.read()
                
                # Save a debug copy
                fallback_filename = f"{base_name}_fallback.xlsx"
                debug_path = os.path.join(os.path.dirname(file_path), fallback_filename)
                with open(debug_path, 'wb') as f:
                    f.write(file_data)
                
                async def send_file_async():
                    file_element = cl.File(
                        name=fallback_filename, 
                        content=file_data, 
                        display="inline", 
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
                    await cl.Message(
                        content=f"Here is your Excel file (fallback method): {fallback_filename}", 
                        elements=[file_element]
                    ).send()
                
                # Run the async function in the sync context
                run_sync(send_file_async())
                
                return f"Excel file '{fallback_filename}' has been sent for download."
            except Exception as fallback_error:
                return f"Error downloading Excel file: {str(fallback_error)}"

class UnmergeExcelCellsTool(BaseTool):
    name: str = "unmerge_excel_cells"
    description: str = "Unmerge all merged cells in the Excel file and fill each unmerged cell with the value from the top-left cell. Input can be 'yes' to confirm."
    
    def _run(self, confirm: str = "yes", run_manager: Optional[CallbackManagerForToolRun] = None) -> str:
        """Unmerge all merged cells in the Excel file."""
        # Get Excel files and current file ID from the session
        excel_files = get_excel_files()
        current_file_id = get_current_file_id()
        
        if current_file_id is None:
            return "No Excel file is currently selected. Use 'load_excel_file' to load a file."
        
        if current_file_id not in excel_files:
            return f"Error: Excel file with ID {current_file_id} not found."
        
        # Get the workbook
        wb = excel_files[current_file_id]['workbook']
        
        # Track statistics for reporting
        total_sheets = 0
        total_merged_ranges = 0
        total_cells_unmerged = 0
        
        # Process each sheet in the workbook
        for sheet in wb:
            total_sheets += 1
            sheet_merged_ranges = 0
            sheet_cells_unmerged = 0
            
            # Iterate through COPY of merged ranges list (since we'll be modifying it)
            for merged_range in list(sheet.merged_cells.ranges):
                sheet_merged_ranges += 1
                
                # Get merged value from top-left cell
                top_left_value = sheet.cell(
                    row=merged_range.min_row, 
                    column=merged_range.min_col
                ).value
                
                # Calculate how many cells are in this range
                cells_in_range = (merged_range.max_row - merged_range.min_row + 1) * (merged_range.max_col - merged_range.min_col + 1)
                sheet_cells_unmerged += cells_in_range - 1  # Subtract 1 for the top-left cell that already has a value
                
                # Unmerge the cells
                sheet.unmerge_cells(str(merged_range))
                
                # Fill value to all cells in original merged area
                for row in range(merged_range.min_row, merged_range.max_row + 1):
                    for col in range(merged_range.min_col, merged_range.max_col + 1):
                        # Skip original top-left cell (already has value)
                        if row == merged_range.min_row and col == merged_range.min_col:
                            continue
                        sheet.cell(row=row, column=col, value=top_left_value)
            
            total_merged_ranges += sheet_merged_ranges
            total_cells_unmerged += sheet_cells_unmerged
        
        # Save the workbook
        file_path = excel_files[current_file_id]['file_path']
        wb.save(file_path)
        
        # Return a summary of what was done
        if total_merged_ranges == 0:
            return "No merged cells found in the workbook."
        
        return (
            f"Successfully unmerged all cells in the workbook:\n"
            f"- Processed {total_sheets} sheets\n"
            f"- Unmerged {total_merged_ranges} merged ranges\n"
            f"- Filled {total_cells_unmerged} cells with values from their respective top-left cells\n"
            f"The workbook has been saved to {file_path}"
        )

class ExcelPythonREPLTool(BaseTool):
    name: str = "execute_openpyxl_code"
    description: str = "Execute openpyxl code on a loaded Excel file. Input should be the Python code to execute."
    
    def _run(self, code: str, run_manager: Optional[CallbackManagerForToolRun] = None) -> str:
        """Execute openpyxl code on a loaded Excel file."""
        print("\n========== STARTING CODE EXECUTION ==========")
        print(f"Original input received: {repr(code)}")
        
        # Simple JSON extraction if needed - only handle the most common case
        if code.strip().startswith("{") and "input" in code:
            try:
                import json
                json_obj = json.loads(code)
                if isinstance(json_obj, dict) and 'input' in json_obj:
                    code = json_obj['input']
                    print(f"Extracted code from JSON: {repr(code)}")
            except Exception as e:
                print(f"JSON parsing failed: {str(e)}")
        
        print(f"\nCode to execute: {repr(code)}")
        
        # Get Excel files and current file ID from Chainlit session
        excel_files = get_excel_files()
        current_file_id = get_current_file_id()
        
        if current_file_id is None:
            return "No Excel file is currently selected. Use 'load_excel_file' to load a file."
        
        if current_file_id not in excel_files:
            return f"Error: Excel file with ID {current_file_id} not found."
        
        # Get the workbook and file path
        wb = excel_files[current_file_id]['workbook']
        file_path = excel_files[current_file_id]['file_path']
        
        # Debug information about the workbook before execution
        try:
            print("========== WORKBOOK STATE BEFORE EXECUTION ==========")
            print(f"Workbook object: {wb}")
            print(f"Workbook is_closed attribute: {getattr(wb, '_Workbook__closed', 'attribute not found')}")
            print(f"Workbook has {len(wb.sheetnames)} sheets: {wb.sheetnames}")
            print(f"Active sheet: {wb.active.title}")
        except Exception as debug_error:
            print(f"Debug info error before execution: {str(debug_error)}")
        
        # Check if the code is trying to load the workbook by filename
        import re
        load_workbook_pattern = r"(?:wb\s*=\s*)?load_workbook\s*\(\s*['\"]([^'\"]+)['\"]\s*\)"
        match = re.search(load_workbook_pattern, code)
        
        if match:
            filename = match.group(1)
            # Replace any load_workbook calls with a comment
            code = re.sub(load_workbook_pattern, "# Using existing wb object instead of load_workbook", code)
            print(f"Detected and removed load_workbook call for file: {filename}")
        
        # Print code after pattern substitution
        print(f"Code after load_workbook substitution: {repr(code)}")
        
        # Check if the workbook is closed and reload if necessary
        try:
            # Simple test to see if the workbook is closed
            test_sheet = wb.active
            sheet_name = test_sheet.title
            print(f"Workbook is open, active sheet: {sheet_name}")
        except Exception as e:
            if "I/O operation on closed file" in str(e):
                print("Workbook is closed, attempting to reload from disk")
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(file_path)
                    # Update the workbook in the session
                    excel_files[current_file_id]['workbook'] = wb
                    cl.user_session.set("excel_files", excel_files)
                    print(f"Successfully reloaded workbook from {file_path}")
                except Exception as reload_error:
                    return f"Error: Workbook is closed and could not be reloaded: {str(reload_error)}"
            else:
                print(f"Unexpected error checking workbook: {str(e)}")
        
        try:
            # Create a local scope with the workbook and file path
            local_vars = {
                "wb": wb, 
                "file_path": file_path,
                "openpyxl": openpyxl,
                "Font": Font,
                "Alignment": Alignment,
                "Border": Border,
                "Side": Side,
                "PatternFill": PatternFill,
                "Color": Color,
                "DifferentialStyle": DifferentialStyle,
                "Rule": Rule,
                "BarChart": BarChart,
                "LineChart": LineChart,
                "PieChart": PieChart,
                "Reference": Reference,
                "Image": Image,
                "get_column_letter": get_column_letter
            }
            
            # Add more debugging prints
            print(f"Starting execution with workbook: {wb}")
            print(f"Workbook has {len(wb.sheetnames)} sheets: {wb.sheetnames}")
            
            # Capture stdout to get print output
            import io
            import sys
            original_stdout = sys.stdout
            captured_output = io.StringIO()
            sys.stdout = captured_output
            
            # Execute the code
            print(f"Executing code with file_path: {file_path}")
            
            # Add more debugging to track execution
            try:
                print("About to execute code...")
                print(f"Code to be executed (final):\n{code}")
                
                # Check for common syntax errors
                try:
                    compile(code, '<string>', 'exec')
                    print("Code compilation check passed")
                except SyntaxError as syntax_err:
                    print(f"SYNTAX ERROR DETECTED: {syntax_err}")
                    print(f"Error on line {syntax_err.lineno}, position {syntax_err.offset}: {syntax_err.text}")
                    raise
                
                # Execute the code
                exec(code, globals(), local_vars)
                print("Code executed successfully")
            except Exception as exec_error:
                print(f"Error during execution: {str(exec_error)}")
                print(f"Error type: {type(exec_error).__name__}")
                print(f"Traceback: {__import__('traceback').format_exc()}")
                raise  # Re-raise to be caught by the outer try/except
            
            # Restore stdout
            sys.stdout = original_stdout
            output = captured_output.getvalue()
            
            # Get the workbook from local_vars - it might have been modified
            wb = local_vars["wb"]
            
            # Debug information about the workbook after execution
            try:
                print("========== WORKBOOK STATE AFTER EXECUTION ==========")
                print(f"Workbook object: {wb}")
                print(f"Workbook is_closed attribute: {getattr(wb, '_Workbook__closed', 'attribute not found')}")
                print(f"Workbook has {len(wb.sheetnames)} sheets: {wb.sheetnames}")
                print(f"Active sheet: {wb.active.title}")
            except Exception as debug_error:
                print(f"Debug info error after execution: {str(debug_error)}")
            
            # Update the workbook in Chainlit session
            excel_files[current_file_id]['workbook'] = wb
            cl.user_session.set("excel_files", excel_files)
            
            # Save the workbook to disk after execution to ensure persistence
            try:
                wb.save(file_path)
                print(f"Automatically saved workbook to {file_path} after code execution")
                
                # Check if the workbook is still open after saving
                try:
                    print("Checking if workbook is still open after saving...")
                    test_sheet = wb.active
                    sheet_name = test_sheet.title
                    print(f"Workbook is still open after saving, active sheet: {sheet_name}")
                except Exception as check_error:
                    print(f"ERROR: Workbook appears to be closed after saving: {str(check_error)}")
            except Exception as save_error:
                print(f"Warning: Could not auto-save workbook after code execution: {str(save_error)}")
            
            # Prepare the result message
            result_message = "Code executed successfully."
            
            # Add captured output if any
            if output.strip():
                # Clean the output (remove debug lines)
                clean_output = output.replace(f"Executing code with file_path: {file_path}", "").strip()
                if clean_output:
                    result_message += f"\n\nOutput:\n{clean_output}"
            
            print("========== CODE EXECUTION COMPLETED SUCCESSFULLY ==========\n")
            return result_message
        except Exception as e:
            print(f"\nERROR executing code: {str(e)}")
            print(f"Error type: {type(e).__name__}")
            print(f"Traceback: {__import__('traceback').format_exc()}")
            print("========== CODE EXECUTION FAILED ==========\n")
            return f"Error executing code: {str(e)}"